"""Normalize all 6 markets into a single unified data file for the map."""
import openpyxl
import json
from pathlib import Path

DATA_DIR = Path("/home/user/workspace/sophi-market-map/data")
OUT = Path("/home/user/workspace/sophi-market-map/src")

MARKETS = {
    "denver":       {"file": "denver_2026_scored_v2.xlsx",   "state": "COLD", "center": [-104.9903, 39.7392]},
    "charlotte":    {"file": "charlotte_2026_scored.xlsx",   "state": "WARM", "center": [-80.8431, 35.2271]},
    "indianapolis": {"file": "indianapolis_2026_scored.xlsx","state": "COLD", "center": [-86.1581, 39.7684]},
    "phoenix":      {"file": "phoenix_2025_scored.xlsx",     "state": "COLD", "center": [-112.0740, 33.4484]},
    "cleveland":    {"file": "cleveland_2026_scored.xlsx",   "state": "COLD", "center": [-81.6944, 41.4993]},
    "louisville":   {"file": "louisville_2026_scored.xlsx",  "state": "COLD", "center": [-85.7585, 38.2527]},
}

# Market display names
MARKET_NAMES = {
    "denver": "Denver, CO",
    "charlotte": "Charlotte, NC",
    "indianapolis": "Indianapolis, IN",
    "phoenix": "Phoenix, AZ",
    "cleveland": "Cleveland, OH",
    "louisville": "Louisville, KY",
}

def normalize_header(h):
    if h is None:
        return ""
    return str(h).strip().lower()

def read_market(market_key, cfg):
    """Extract accounts + scoring for a single market."""
    wb = openpyxl.load_workbook(DATA_DIR / cfg["file"], data_only=True)
    accounts_ws = wb["Accounts"]

    # Find header row (first row with "Account Type")
    headers_row_idx = None
    for i, row in enumerate(accounts_ws.iter_rows(max_row=5, values_only=True), 1):
        if row and any(normalize_header(c) == "account type" for c in row if c):
            headers_row_idx = i
            break
    if headers_row_idx is None:
        headers_row_idx = 1

    # Read headers
    headers = [normalize_header(c.value) for c in accounts_ws[headers_row_idx]]

    # Build column map
    def col(name_options):
        for i, h in enumerate(headers):
            for opt in name_options:
                if opt in h:
                    return i
        return None

    c_name    = col(["account"]) if "account" in headers else 0
    # First column is account name — fallback to index 0
    if c_name is None or headers[0] in ("", "account"):
        c_name = 0
    c_type    = col(["account type"])
    c_addr    = col(["address", "area"])
    c_downtown= col(["downtown", "uptown"])
    c_phone   = col(["phone number"])
    c_email   = col(["email"])
    c_url     = col(["url"])
    c_self    = col(["self parking rate"])
    c_valet   = col(["valet rate"])
    c_rooms   = col(["rooms", "room count", "beds"])
    c_gm      = col(["gm name", "poc name"])
    c_mgmt    = col(["management group"])
    c_garage  = col(["garage operator"])
    c_valetop = col(["valet operator"])
    c_notes   = col(["location notes"])
    c_source  = col(["sourcing notes"])

    # Read WAS scoring
    was = {}
    if "WAS_Scoring_v2" in wb.sheetnames:
        ws = wb["WAS_Scoring_v2"]
        wh = [normalize_header(c.value) for c in ws[1]]
        def wc(n):
            for i, h in enumerate(wh):
                if n in h:
                    return i
            return None
        cw_name = 0
        cw_tam  = wc("tam")
        cw_was  = wc("was final")
        if cw_was is None:
            cw_was = wc("was raw")
        cw_tier = wc("tier")
        cw_y1r  = wc("y1 rev")
        cw_y5r  = wc("y5 rev")
        cw_fit  = wc("fit")
        cw_size = wc("size")
        cw_own  = wc("owner")
        cw_adj  = wc("adj")
        cw_rel  = wc("rel tag")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[cw_name]:
                continue
            key = str(row[cw_name]).strip().lower()
            was[key] = {
                "tam":    row[cw_tam] if cw_tam is not None else None,
                "was":    row[cw_was] if cw_was is not None else None,
                "tier":   row[cw_tier] if cw_tier is not None else None,
                "y1_rev": row[cw_y1r] if cw_y1r is not None else None,
                "y5_rev": row[cw_y5r] if cw_y5r is not None else None,
                "fit":    row[cw_fit] if cw_fit is not None else None,
                "size":   row[cw_size] if cw_size is not None else None,
                "owner":  row[cw_own] if cw_own is not None else None,
                "adj":    row[cw_adj] if cw_adj is not None else None,
                "rel_tag":row[cw_rel] if cw_rel is not None else None,
            }

    accounts = []
    for row in accounts_ws.iter_rows(min_row=headers_row_idx+1, values_only=True):
        name = row[c_name] if c_name < len(row) else None
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()
        acct_type = row[c_type] if c_type is not None and c_type < len(row) else None
        if not acct_type:
            continue
        acct_type = str(acct_type).strip()

        # skip empty/summary rows
        if acct_type.lower() in ("", "none", "total"):
            continue

        def val(i):
            if i is None or i >= len(row):
                return None
            v = row[i]
            if v is None:
                return None
            s = str(v).strip()
            return s if s else None

        was_data = was.get(name_str.lower(), {})

        account = {
            "name": name_str,
            "type": acct_type,
            "address": val(c_addr),
            "downtown": val(c_downtown),
            "phone": val(c_phone),
            "email": val(c_email),
            "url": val(c_url),
            "self_park_rate": val(c_self),
            "valet_rate": val(c_valet),
            "rooms": val(c_rooms),
            "gm": val(c_gm),
            "management": val(c_mgmt),
            "garage_operator": val(c_garage),
            "valet_operator": val(c_valetop),
            "location_notes": val(c_notes),
            "sourcing_notes": val(c_source),
            # scoring
            "tam": was_data.get("tam"),
            "was_score": was_data.get("was"),
            "tier": was_data.get("tier"),
            "y1_rev": was_data.get("y1_rev"),
            "y5_rev": was_data.get("y5_rev"),
            "fit": was_data.get("fit"),
            "size_score": was_data.get("size"),
            "owner_base": was_data.get("owner"),
            "addressability": was_data.get("adj"),  # note: misused field
            "relationship": was_data.get("rel_tag"),
        }
        accounts.append(account)

    wb.close()
    return accounts

# 6-market rollup summary
def read_rollup():
    wb = openpyxl.load_workbook(DATA_DIR / "sophi_6market_rollup.xlsx", data_only=True)
    ws = wb["Master_Summary"]
    summaries = {}
    for row in ws.iter_rows(min_row=6, max_row=11, values_only=True):
        if row[0] and "TOTAL" not in str(row[0]):
            summaries[str(row[0]).lower()] = {
                "state": row[1],
                "n_accounts": row[2],
                "tam": row[3],
                "y1_som": row[4],
                "y2_som": row[5],
                "y3_som": row[6],
                "y4_som": row[7],
                "y5_som": row[8],
            }
    # totals
    totals = None
    for row in ws.iter_rows(min_row=11, max_row=11, values_only=True):
        totals = {
            "n_accounts": row[2],
            "tam": row[3],
            "y5_som": row[8],
        }
    wb.close()
    return summaries, totals

if __name__ == "__main__":
    all_data = {"markets": {}, "meta": {}}

    for mkey, cfg in MARKETS.items():
        print(f"Processing {mkey}...")
        accounts = read_market(mkey, cfg)
        all_data["markets"][mkey] = {
            "name": MARKET_NAMES[mkey],
            "state": cfg["state"],
            "center": cfg["center"],
            "accounts": accounts,
        }
        print(f"  {len(accounts)} accounts loaded")

    summaries, totals = read_rollup()
    for mkey in MARKETS:
        if mkey in summaries:
            all_data["markets"][mkey]["summary"] = summaries[mkey]
    all_data["meta"]["totals"] = totals
    all_data["meta"]["methodology"] = {
        "weights": {"fit": 0.25, "size": 0.25, "ownership": 0.20, "addressability": 0.15, "adjacency": 0.15},
        "tiers": {
            "A — Beachhead": "Top priority targets with highest WAS score",
            "B — Core": "Core growth targets - major revenue contributors",
            "C — Opportunistic": "Opportunistic accounts with moderate fit",
            "D — Deprioritized": "Lower priority, long-term prospects",
        },
        "cold_curves": {"A": "40% → 95%", "B": "8% → 75%", "C": "2% → 40%", "D": "0% → 10%"},
        "warm_curves": {"A": "60% → 95%", "B": "25% → 80%", "C": "10% → 45%", "D": "0% → 15%"},
    }

    out_path = OUT / "accounts.json"
    with open(out_path, "w") as f:
        json.dump(all_data, f, indent=2, default=str)
    print(f"\n✓ Wrote {out_path}")

    # Print summary
    total_accts = sum(len(all_data["markets"][m]["accounts"]) for m in MARKETS)
    print(f"Total accounts: {total_accts}")
    for mkey in MARKETS:
        print(f"  {mkey}: {len(all_data['markets'][mkey]['accounts'])}")
