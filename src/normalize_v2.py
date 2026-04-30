#!/usr/bin/env python3
"""
Normalize 6 markets from v5/v7 XLSX files to unified v2-methodology JSON.

Truth-source strategy (v3 of normalizer):
- SOM_v5 sheet IS THE CANONICAL ACCOUNT LIST for 5 markets (Charlotte/Phoenix/Cleveland/
  Louisville/Denver). It contains both in-SAM and out-of-SAM accounts with `Pool` tags.
- For Indianapolis (v7), SOM_v5 only contains in-SAM (14). Out-of-SAM accounts (14)
  are reconstructed from Towne_Reclass_Log + TAM_Conservative + a curated mapping.
- TAM_Conservative + Accounts sheets enrich each SOM_v5 account with contact/address/etc.
- Rollup (Portfolio_Rollup, Pool_Structure, Sensitivity) gives portfolio truth numbers.

Outputs: /home/user/workspace/sophi-market-map/src/accounts_v2.json
"""
import json
import os
import re
from openpyxl import load_workbook

BASE = '/home/user/workspace/past_session_contexts/sessions/2026-04-27_2026-05-03/07d6c972/ai_outputs/'
OUT  = '/home/user/workspace/sophi-market-map/src/accounts_v2.json'

MARKETS = [
    ('charlotte',    'Charlotte, NC',    'WARM', 'charlotte_2026_som_v5.xlsx'),
    ('phoenix',      'Phoenix, AZ',      'COLD', 'phoenix_2025_som_v5.xlsx'),
    ('cleveland',    'Cleveland, OH',    'COLD', 'cleveland_2026_som_v5.xlsx'),
    ('louisville',   'Louisville, KY',   'COLD', 'louisville_2026_som_v5.xlsx'),
    ('denver',       'Denver, CO',       'COLD', 'denver_2026_som_v5.xlsx'),
    ('indianapolis', 'Indianapolis, IN', 'COLD', 'indianapolis_2026_som_v5.xlsx'),
]

# Indianapolis manual out-of-SAM list (14 accounts to supplement SOM_v5's 14 in-SAM)
# Derived from Towne_Reclass_Log + TAM_Conservative + Pool_Structure rollup truth
INDY_OUT_OF_SAM = [
    # Enterprise (6) - from Towne_Reclass_Log
    {'name': 'JW Marriott Indianapolis',                'pool': 'enterprise',    'operator': 'White Lodging / Towne Park'},
    {'name': 'Marriott Indianapolis Downtown',          'pool': 'enterprise',    'operator': 'White Lodging / Towne Park'},
    {'name': 'Hyatt Regency Indianapolis',              'pool': 'enterprise',    'operator': 'Towne Park'},
    {'name': 'The Westin Indianapolis',                 'pool': 'enterprise',    'operator': 'Towne Park'},
    {'name': 'Hilton Indianapolis Hotel & Suites',      'pool': 'enterprise',    'operator': 'Towne Park'},
    {'name': 'Crowne Plaza Indianapolis Downtown–Union Station', 'pool': 'enterprise', 'operator': 'Towne Park'},
    # Partnership (2)
    {'name': 'Le Méridien Indianapolis',                'pool': 'partnership',   'operator': 'PMC'},
    {'name': 'Bottleworks Hotel',                       'pool': 'partnership',   'operator': 'PMC'},
    # Extended-stay (1)
    {'name': 'Homewood Suites by Hilton Indianapolis Downtown', 'pool': 'extended_stay', 'operator': 'PMC'},
    # Micro (5) - lowest-TAM downtown/suburban hotels & restaurants
    {'name': 'Hotel Carmichael, Autograph Collection',  'pool': 'micro',         'operator': ''},
    {'name': "Mo's A Place for Steaks",                 'pool': 'micro',         'operator': ''},
    {'name': 'Provision (at JW Marriott)',              'pool': 'micro',         'operator': ''},
    {'name': 'Bluebeard',                               'pool': 'micro',         'operator': ''},
    {'name': 'Restaurant at Hotel Carmichael',          'pool': 'micro',         'operator': ''},
]


def norm(s):
    if s is None: return ''
    return re.sub(r'\s+', ' ', str(s)).strip()


def safe_float(v):
    if v is None or v == '': return 0
    try: return float(v)
    except (ValueError, TypeError): return 0


def read_sheet_rows(wb, name):
    if name not in wb.sheetnames: return []
    return list(wb[name].iter_rows(values_only=True))


def parse_som_v5(rows):
    """Returns ordered list of accounts from SOM_v5 sheet with pool tags."""
    # Find header row (Rank, Account, ..., Pool, ...)
    hdr_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == 'Rank':
            hdr_idx = i
            break
    if hdr_idx is None: return []
    header = [norm(c) for c in rows[hdr_idx]]

    def col(name):
        return header.index(name) if name in header else None

    i_rank  = col('Rank')
    i_acct  = col('Account')
    i_op    = col('Operator')
    i_mgmt  = col('Mgmt')
    i_area  = col('Area')
    i_tam   = col('TAM')
    i_sam   = col('SAM Contrib')
    i_was_a = col('WAS Adj')
    i_was   = col('WAS')
    i_was_b = col('WAS Base')
    i_boost = col('Boost')
    i_pool  = col('Pool')
    i_sy    = col('Sign Yr')
    i_crv   = col('Curve')
    i_y1    = col('Y1')
    i_y5    = col('Y5')

    out = []
    for row in rows[hdr_idx + 1:]:
        if not row: continue
        acct = norm(row[i_acct]) if i_acct is not None else ''
        if not acct: continue
        # Skip section headers and totals
        if acct.startswith('—'): continue
        if acct.upper().startswith('TOTAL'): continue
        if 'SUBTOTAL' in acct.upper(): continue
        if acct.upper().startswith('ORGANIC'): continue
        if 'M&A' in acct.upper() and 'ABSORPTION' in acct.upper(): continue

        tam = safe_float(row[i_tam]) if i_tam is not None else 0
        sam_c = safe_float(row[i_sam]) if i_sam is not None else 0
        # WAS: prefer Adj > base > raw 'WAS'
        was = None
        if i_was_a is not None and row[i_was_a] is not None:
            was = safe_float(row[i_was_a])
        elif i_was is not None and row[i_was] is not None:
            was = safe_float(row[i_was])
        elif i_was_b is not None and row[i_was_b] is not None:
            was = safe_float(row[i_was_b])

        was_base = safe_float(row[i_was_b]) if i_was_b is not None and row[i_was_b] is not None else None
        boost = safe_float(row[i_boost]) if i_boost is not None and row[i_boost] is not None else None

        out.append({
            'name': acct,
            'rank': norm(row[i_rank]) if i_rank is not None else None,
            'operator_som': norm(row[i_op]) if i_op is not None else '',
            'mgmt_som': norm(row[i_mgmt]) if i_mgmt is not None else '',
            'area_som': norm(row[i_area]) if i_area is not None else '',
            'tam_som': tam,
            'sam_contrib': sam_c,
            'was': was,
            'was_base': was_base,
            'was_boost': boost,
            'pool_raw': norm(row[i_pool]) if i_pool is not None else '',
            'sign_yr': norm(row[i_sy]) if i_sy is not None else None,
            'curve': norm(row[i_crv]) if i_crv is not None else '',
            'y1': safe_float(row[i_y1]) if i_y1 is not None else 0,
            'y2': safe_float(row[i_y1+1]) if i_y1 is not None else 0,
            'y3': safe_float(row[i_y1+2]) if i_y1 is not None else 0,
            'y4': safe_float(row[i_y1+3]) if i_y1 is not None else 0,
            'y5': safe_float(row[i_y5]) if i_y5 is not None else 0,
        })
    return out


def parse_tam_conservative(rows):
    """Returns dict by lowercase name → TAM enrichment."""
    if not rows: return {}
    header = [norm(c) for c in rows[0]]
    def col(n): return header.index(n) if n in header else None
    i_acct = col('Account'); i_type = col('Account Type'); i_class = col('TAM Class')
    i_rooms = col('Rooms/Beds'); i_seats = col('Seats')
    i_occ = col('Occupancy %'); i_turn = col('Turnover')
    i_conv = col('Valet Conv %'); i_rate = col('Valet Rate')
    i_tam = col('TAM'); i_status = col('TAM Status'); i_notes = col('Notes / Source')

    out = {}
    for row in rows[1:]:
        if not row or not row[i_acct]: continue
        acct = norm(row[i_acct])
        if not acct or acct.startswith('TOTAL'): continue
        out[acct.lower()] = {
            'name_tam': acct,
            'type': norm(row[i_type]) if i_type is not None else '',
            'tam_class': norm(row[i_class]) if i_class is not None else '',
            'rooms': row[i_rooms] if i_rooms is not None else None,
            'seats': row[i_seats] if i_seats is not None else None,
            'occupancy': row[i_occ] if i_occ is not None else None,
            'turnover': row[i_turn] if i_turn is not None else None,
            'valet_conv': row[i_conv] if i_conv is not None else None,
            'valet_rate': row[i_rate] if i_rate is not None else None,
            'tam_calc': safe_float(row[i_tam]) if i_tam is not None else 0,
            'tam_status': norm(row[i_status]) if i_status is not None else '',
            'tam_notes': norm(row[i_notes]) if i_notes is not None else '',
        }
    return out


def parse_accounts(rows):
    """Returns dict by lowercase name → contact/address enrichment."""
    if not rows: return {}
    # Find header row (within first 3 rows) - prefer 'Address' header, else any 'Account Type'
    header = None
    header_idx = 0
    for i, row in enumerate(rows[:3]):
        if row and any(c and 'Address' in str(c) for c in row):
            header = [norm(c) for c in row]
            header_idx = i
            break
    # Fallback: first row with 'Account Type' (Phoenix has unlabeled Address column)
    if header is None:
        for i, row in enumerate(rows[:3]):
            if row and any(c and 'Account Type' in str(c) for c in row):
                header = [norm(c) for c in row]
                header_idx = i
                break
    if header is None: return {}

    def col(name, alts=None):
        cands = [name] + (alts or [])
        for c in cands:
            for j, h in enumerate(header):
                if h.lower() == c.lower(): return j
        return None

    i_addr = col('Address')
    # Phoenix: col 3 is unlabeled but contains Address
    if i_addr is None and len(header) > 3 and not header[3]:
        # Verify by sniffing first data row
        for row in rows[header_idx+1:header_idx+5]:
            if row and len(row) > 3 and row[3] and any(c in str(row[3]) for c in [',', 'St', 'Ave', 'Dr', 'Rd', 'Blvd']):
                i_addr = 3
                break
    i_type = col('Account Type', ['Type'])
    i_dt   = col('Uptown (Yes or No)', ['Downtown', 'Uptown', 'Area'])
    i_phone= col('Phone Number', ['Phone'])
    i_email= col('Email')
    i_url  = col('URL', ['Website'])
    i_self = col('Self Parking Rate', ['Self-Park Rate'])
    i_valet= col('Valet Rate')
    i_rooms= col('# of Rooms', ['Rooms', '# Rooms', 'Room Count'])
    i_gm   = col('GM Name', ['GM', 'POC Name'])
    i_role = col('POC Role')
    i_mgmt = col('Management Group', ['Management'])
    i_gop  = col('Garage Operator', ['Garage Operator '])
    i_vop  = col('Valet Operator')
    i_src  = col('Sourcing Notes')
    i_loc  = col('Location Notes')

    out = {}
    for row in rows[header_idx + 1:]:
        if not row or not row[0]: continue
        name = norm(row[0])
        if not name: continue
        out[name.lower()] = {
            'name_acct': name,
            'address': norm(row[i_addr]) if i_addr is not None and row[i_addr] else '',
            'type_acct': norm(row[i_type]) if i_type is not None and row[i_type] else '',
            'area_acct': norm(row[i_dt]) if i_dt is not None and row[i_dt] else '',
            'phone': norm(row[i_phone]) if i_phone is not None and row[i_phone] else '',
            'email': norm(row[i_email]) if i_email is not None and row[i_email] else '',
            'url': norm(row[i_url]) if i_url is not None and row[i_url] else '',
            'self_park_rate': row[i_self] if i_self is not None else None,
            'valet_rate_acct': row[i_valet] if i_valet is not None else None,
            'rooms_acct': row[i_rooms] if i_rooms is not None else None,
            'gm': norm(row[i_gm]) if i_gm is not None and row[i_gm] else '',
            'gm_role': norm(row[i_role]) if i_role is not None and row[i_role] else '',
            'management': norm(row[i_mgmt]) if i_mgmt is not None and row[i_mgmt] else '',
            'garage_operator': norm(row[i_gop]) if i_gop is not None and row[i_gop] else '',
            'valet_operator': norm(row[i_vop]) if i_vop is not None and row[i_vop] else '',
            'sourcing_notes': norm(row[i_src]) if i_src is not None and row[i_src] else '',
            'location_notes': norm(row[i_loc]) if i_loc is not None and row[i_loc] else '',
        }
    return out


def normalize_pool_tag(raw_pool):
    """Map raw SOM_v5 pool string to canonical tag."""
    if not raw_pool: return None
    p = raw_pool.lower()
    if 'anchor' in p: return 'anchor'
    if 'absorbed' in p or ('m&a' in p and 'absorpt' not in p): return 'ma_sam'
    if 'cold sam' in p: return 'cold_sam'
    if 'partnership' in p or 'preferred' in p: return 'partnership'
    if 'enterprise' in p: return 'enterprise'
    if 'product-fit' in p or 'extended' in p: return 'extended_stay'
    if 'micro' in p or 'sub-floor' in p or 'sub floor' in p: return 'micro'
    return None


def find_match(name, lookup_dict):
    """Fuzzy match name against lookup dict (lowercase keys)."""
    n = name.lower().strip()
    if n in lookup_dict: return lookup_dict[n]
    # Try strip punctuation
    n2 = re.sub(r'[^\w\s]', '', n)
    for k, v in lookup_dict.items():
        k2 = re.sub(r'[^\w\s]', '', k)
        if n2 == k2: return v
    # Try first 25 chars
    for k, v in lookup_dict.items():
        if n[:25] == k[:25] and len(n) > 15: return v
    return {}


def build_v7_metadata(market, name, pool):
    """Indianapolis v7 hometown displaced / WAS boost / M&A absorption flags."""
    if market != 'indianapolis': return None
    n = name.lower()
    if pool == 'ma_sam':
        return 'ma_absorption'
    if pool == 'cold_sam':
        if 'omni severin' in n or 'conrad indianapolis' in n: return 'hometown_displaced'
        if 'capital grille' in n: return 'hometown_displaced'  # co-located w/ Conrad
        if 'hilton garden inn indianapolis' in n: return 'hometown_was_boost'
    return None


def main():
    portfolio = {'markets': {}, 'meta': {
        'methodology': 'v2 (April 2026) — decoupled WAS/SAM/SOM with 4 structural pools + 50% TAM ceiling. Indianapolis v7 (hometown advantage + Y2 Elite M&A).',
    }}

    grand_total_accts = 0

    for key, name, state, fn in MARKETS:
        path = os.path.join(BASE, fn)
        wb = load_workbook(path, data_only=True)

        som_list = parse_som_v5(read_sheet_rows(wb, 'SOM_v5'))
        tamc     = parse_tam_conservative(read_sheet_rows(wb, 'TAM_Conservative'))
        accts    = parse_accounts(read_sheet_rows(wb, 'Accounts'))

        unified = []

        # Process SOM_v5 accounts (the canonical list for 5/6 markets)
        for som_d in som_list:
            display = som_d['name']
            tam_d  = find_match(display, tamc)
            acct_d = find_match(display, accts)

            pool = normalize_pool_tag(som_d.get('pool_raw'))
            if not pool:
                # Indy v7 has rows missing Pool tag (M&A absorbed) — fallback by raw pool
                p_raw = som_d.get('pool_raw', '').lower()
                if 'absorbed' in p_raw or som_d.get('rank', '').startswith('M'):
                    pool = 'ma_sam'
                else:
                    pool = 'cold_sam'  # default for in-SAM rows in SOM_v5

            in_sam = pool in ('anchor', 'cold_sam', 'ma_sam')

            tam = som_d['tam_som'] or tam_d.get('tam_calc', 0) or 0
            atype = (acct_d.get('type_acct') or tam_d.get('type') or '').strip() or 'Other'
            valet_op = som_d.get('operator_som') or acct_d.get('valet_operator') or ''
            mgmt = som_d.get('mgmt_som') or acct_d.get('management') or ''
            area = som_d.get('area_som') or acct_d.get('area_acct') or ''

            rec = {
                'name': display,
                'market': key,
                'type': atype,
                'tam': tam,
                'sam_contrib': som_d['sam_contrib'],
                'pool': pool,
                'in_sam': in_sam,
                'rank': som_d.get('rank'),
                'was': som_d.get('was'),
                'was_base': som_d.get('was_base'),
                'was_boost': som_d.get('was_boost'),
                'sign_yr': som_d.get('sign_yr'),
                'curve': som_d.get('curve'),
                'y1': som_d['y1'], 'y2': som_d['y2'], 'y3': som_d['y3'],
                'y4': som_d['y4'], 'y5': som_d['y5'],
                'address': acct_d.get('address', ''),
                'area': area,
                'phone': acct_d.get('phone', ''),
                'email': acct_d.get('email', ''),
                'url': acct_d.get('url', ''),
                'self_park_rate': acct_d.get('self_park_rate'),
                'valet_rate': acct_d.get('valet_rate_acct') or tam_d.get('valet_rate'),
                'rooms': acct_d.get('rooms_acct') or tam_d.get('rooms'),
                'seats': tam_d.get('seats'),
                'occupancy': tam_d.get('occupancy'),
                'turnover': tam_d.get('turnover'),
                'valet_conv': tam_d.get('valet_conv'),
                'gm': acct_d.get('gm', ''),
                'gm_role': acct_d.get('gm_role', ''),
                'management': mgmt,
                'garage_operator': acct_d.get('garage_operator', ''),
                'valet_operator': valet_op,
                'sourcing_notes': acct_d.get('sourcing_notes', ''),
                'location_notes': acct_d.get('location_notes', ''),
                'tam_class': tam_d.get('tam_class', ''),
                'tam_status': tam_d.get('tam_status', ''),
                'tam_notes': tam_d.get('tam_notes', ''),
                'pool_raw': som_d.get('pool_raw', ''),
                'v7_layer': build_v7_metadata(key, display, pool),
            }
            unified.append(rec)

        # Indianapolis: append manual out-of-SAM list
        if key == 'indianapolis':
            for entry in INDY_OUT_OF_SAM:
                display = entry['name']
                tam_d  = find_match(display, tamc)
                acct_d = find_match(display, accts)
                tam = tam_d.get('tam_calc', 0) or 0
                atype = (acct_d.get('type_acct') or tam_d.get('type') or '').strip() or 'Hotel'

                rec = {
                    'name': display,
                    'market': key,
                    'type': atype,
                    'tam': tam,
                    'sam_contrib': 0,
                    'pool': entry['pool'],
                    'in_sam': False,
                    'rank': '—',
                    'was': None, 'was_base': None, 'was_boost': None,
                    'sign_yr': 'N/A', 'curve': 'Not in SAM',
                    'y1': 0, 'y2': 0, 'y3': 0, 'y4': 0, 'y5': 0,
                    'address': acct_d.get('address', ''),
                    'area': acct_d.get('area_acct', ''),
                    'phone': acct_d.get('phone', ''),
                    'email': acct_d.get('email', ''),
                    'url': acct_d.get('url', ''),
                    'self_park_rate': acct_d.get('self_park_rate'),
                    'valet_rate': acct_d.get('valet_rate_acct') or tam_d.get('valet_rate'),
                    'rooms': acct_d.get('rooms_acct') or tam_d.get('rooms'),
                    'seats': tam_d.get('seats'),
                    'occupancy': tam_d.get('occupancy'),
                    'turnover': tam_d.get('turnover'),
                    'valet_conv': tam_d.get('valet_conv'),
                    'gm': acct_d.get('gm', ''),
                    'gm_role': acct_d.get('gm_role', ''),
                    'management': acct_d.get('management', ''),
                    'garage_operator': acct_d.get('garage_operator', ''),
                    'valet_operator': entry['operator'] or acct_d.get('valet_operator', ''),
                    'sourcing_notes': acct_d.get('sourcing_notes', ''),
                    'location_notes': acct_d.get('location_notes', ''),
                    'tam_class': tam_d.get('tam_class', ''),
                    'tam_status': tam_d.get('tam_status', ''),
                    'tam_notes': tam_d.get('tam_notes', ''),
                    'pool_raw': entry['pool'],
                    'v7_layer': None,
                }
                unified.append(rec)

        # Pool counts and aggregates
        pool_counts = {p: 0 for p in ['anchor','cold_sam','ma_sam','partnership','enterprise','extended_stay','micro']}
        pool_tam = {p: 0 for p in pool_counts}
        for r in unified:
            pool_counts[r['pool']] = pool_counts.get(r['pool'], 0) + 1
            pool_tam[r['pool']] = pool_tam.get(r['pool'], 0) + r['tam']

        market_tam = sum(r['tam'] for r in unified)
        market_sam_actual = sum(r['sam_contrib'] for r in unified if r['in_sam'])
        market_y1 = sum(r['y1'] for r in unified)
        market_y5 = sum(r['y5'] for r in unified)

        portfolio['markets'][key] = {
            'name': name,
            'state': state,
            'accounts': unified,
            'pool_counts': pool_counts,
            'pool_tam': pool_tam,
            'tier_counts': {
                'A': sum(1 for r in unified if r['was'] and r['was'] >= 4.0),
                'B': sum(1 for r in unified if r['was'] and 3.4 <= r['was'] < 4.0),
                'C': sum(1 for r in unified if r['was'] and 2.8 <= r['was'] < 3.4),
                'D': sum(1 for r in unified if r['was'] and r['was'] < 2.8),
            },
            'summary': {
                'tam': market_tam, 'sam': market_sam_actual,
                'y1_som': market_y1, 'y5_som': market_y5,
                'n_accounts': len(unified),
                'n_in_sam': sum(1 for r in unified if r['in_sam']),
                'state': state,
            }
        }
        grand_total_accts += len(unified)
        print(f'[{key}] {len(unified)} accts | TAM ${market_tam/1e6:.1f}M | SAM ${market_sam_actual/1e6:.1f}M | Y5 ${market_y5/1e6:.2f}M')
        for pk in ['anchor','cold_sam','ma_sam','partnership','enterprise','extended_stay','micro']:
            if pool_counts[pk]:
                print(f'    {pk:>16}: {pool_counts[pk]:>3} accts  ${pool_tam[pk]/1e6:.2f}M')

    # Pull rollup truth numbers
    rollup_path = os.path.join(BASE, 'sophi_6market_rollup.xlsx')
    rwb = load_workbook(rollup_path, data_only=True)
    rrows = list(rwb['Portfolio_Rollup'].iter_rows(values_only=True))
    rmap = {'Charlotte':'charlotte','Phoenix':'phoenix','Cleveland':'cleveland',
            'Louisville':'louisville','Denver':'denver','Indianapolis':'indianapolis'}
    for row in rrows:
        if not row or not row[0]: continue
        market_name = norm(row[0])
        for k_in, k_out in rmap.items():
            if market_name.startswith(k_in):
                m = portfolio['markets'][k_out]
                m['rollup'] = {
                    'market_type': row[1], 'scope': row[2], 'n_accts': row[3],
                    'tam': row[4], 'sam': row[5], 'sam_tam_ratio': row[6],
                    'y1': row[7], 'y2': row[8], 'y3': row[9], 'y4': row[10], 'y5': row[11],
                    'y5_tam_ratio': row[12], 'y5_sam_ratio': row[13], 'notes': row[14],
                }
                # Rollup truth overrides our calculated summary
                m['summary']['tam'] = row[4]
                m['summary']['sam'] = row[5]
                m['summary']['y1_som'] = row[7]
                m['summary']['y5_som'] = row[11]
                m['summary']['sam_tam_ratio'] = row[6]
                m['summary']['y5_tam_ratio'] = row[12]
                m['summary']['y5_sam_ratio'] = row[13]

    # Pool_Structure rollup numbers (the per-pool TAM truth for headers)
    psr = list(rwb['Pool_Structure'].iter_rows(values_only=True))
    for row in psr:
        if not row or not row[0]: continue
        mname = norm(row[0])
        if mname not in rmap: continue
        m = portfolio['markets'][rmap[mname]]
        m['pool_structure_rollup'] = {
            'anchor':       {'count': row[1] or 0, 'tam': row[2] or 0},
            'cold_sam':     {'count': row[3] or 0, 'tam': row[4] or 0},
            'ma_sam':       {'count': row[5] or 0, 'tam': row[6] or 0},
            'partnership':  {'count': row[7] or 0, 'tam': row[8] or 0},
            'enterprise':   {'count': row[9] or 0, 'tam': row[10] or 0},
            'extended_stay':{'count': row[11] or 0, 'tam': row[12] or 0},
            'micro':        {'count': row[13] or 0, 'tam': row[14] or 0},
            'total':        {'count': row[15] or 0, 'tam': row[16] or 0},
        }

    # Sensitivity scenarios
    srows = list(rwb['Sensitivity'].iter_rows(values_only=True))
    sensitivity = []
    for row in srows[4:]:
        if not row or not row[0]: continue
        sensitivity.append({
            'scenario': norm(row[0]), 'tam': row[1], 'sam': row[2],
            'y1': row[3], 'y2': row[4], 'y3': row[5], 'y4': row[6], 'y5': row[7],
            'y5_tam': row[8], 'y5_sam': row[9],
        })
    portfolio['meta']['sensitivity'] = sensitivity

    # LP_Headlines
    lp_rows = list(rwb['LP_Headlines'].iter_rows(values_only=True))
    headlines = []
    for row in lp_rows[3:]:
        if not row or not row[0]: continue
        headlines.append({'metric': norm(row[0]), 'value': row[1], 'note': norm(row[2])})
    portfolio['meta']['lp_headlines'] = headlines

    # Portfolio totals from rollup
    portfolio['meta']['total_tam'] = sum(m['summary']['tam'] for m in portfolio['markets'].values())
    portfolio['meta']['total_sam'] = sum(m['summary']['sam'] for m in portfolio['markets'].values())
    portfolio['meta']['total_y1_som'] = sum(m['summary']['y1_som'] for m in portfolio['markets'].values())
    portfolio['meta']['total_y5_som'] = sum(m['summary']['y5_som'] for m in portfolio['markets'].values())
    portfolio['meta']['total_accounts'] = grand_total_accts
    portfolio['meta']['total_in_sam'] = sum(m['summary']['n_in_sam'] for m in portfolio['markets'].values())

    print('\n=== PORTFOLIO ===')
    print(f"  TAM:    ${portfolio['meta']['total_tam']/1e6:.1f}M")
    print(f"  SAM:    ${portfolio['meta']['total_sam']/1e6:.1f}M  ({portfolio['meta']['total_sam']/portfolio['meta']['total_tam']*100:.1f}% of TAM)")
    print(f"  Y1 SOM: ${portfolio['meta']['total_y1_som']/1e6:.1f}M")
    print(f"  Y5 SOM: ${portfolio['meta']['total_y5_som']/1e6:.1f}M  ({portfolio['meta']['total_y5_som']/portfolio['meta']['total_tam']*100:.1f}% of TAM)")
    print(f"  Accts:  {portfolio['meta']['total_accounts']} ({portfolio['meta']['total_in_sam']} in SAM)")

    # Validation against target pool counts
    target_pool_counts = {
        'charlotte':    {'anchor':4,'cold_sam':20,'ma_sam':0,'partnership':6,'enterprise':1,'extended_stay':4,'micro':4},
        'phoenix':      {'anchor':0,'cold_sam':19,'ma_sam':0,'partnership':4,'enterprise':12,'extended_stay':2,'micro':0},
        'cleveland':    {'anchor':0,'cold_sam':8, 'ma_sam':0,'partnership':0,'enterprise':7,'extended_stay':0,'micro':10},
        'louisville':   {'anchor':0,'cold_sam':7, 'ma_sam':0,'partnership':1,'enterprise':6,'extended_stay':0,'micro':11},
        'denver':       {'anchor':0,'cold_sam':21,'ma_sam':0,'partnership':4,'enterprise':17,'extended_stay':5,'micro':2},
        'indianapolis': {'anchor':0,'cold_sam':9, 'ma_sam':5,'partnership':2,'enterprise':6,'extended_stay':1,'micro':5},
    }
    print('\n=== VALIDATION ===')
    all_ok = True
    for mkt, tgt in target_pool_counts.items():
        cur = portfolio['markets'][mkt]['pool_counts']
        ok = all(cur.get(p, 0) == tgt[p] for p in tgt)
        flag = ' ✓ OK' if ok else ' ✗ MISMATCH'
        print(f'  {mkt:14s} {flag}')
        if not ok:
            all_ok = False
            for p, v in tgt.items():
                if cur.get(p, 0) != v:
                    print(f'    {p}: got {cur.get(p,0)}, want {v}')
    print(f"\n{'✓ ALL POOL COUNTS MATCH' if all_ok else '✗ POOL COUNT MISMATCHES — REVIEW'}")

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(portfolio, f, default=str, ensure_ascii=False, indent=2)
    print(f'\nwrote {OUT} ({os.path.getsize(OUT):,} bytes)')


if __name__ == '__main__':
    main()
